import openpyxl
import random

# ブック、シートを開く
wb = openpyxl.load_workbook("2020年度_健康管理チェックシート2021版.xlsx")
ws = wb["11月"]

b_temp = [36.0, 36.1, 36.2, 36.3, 36.4, 36.5, 36.6, 36.7, 36.8, 35.9, 35.8]
sel_min = [0, 5, 10, 15, 20, 25, 30, 35, 40, 45, 50, 55]
sel_hour_mor = [6, 7, 8]
sel_hour_eve = [19, 20, 21, 22]

start_day = int(input())
for j in range(7):
    for i in range(6):
        if ((i == 0) or (i == 3)):
            select_item = random.choice(b_temp)
        elif (i == 1):
            select_item = random.choice(sel_hour_mor)
        elif (i == 4):
            select_item = random.choice(sel_hour_eve)
        else:
            select_item = random.choice(sel_min)
        ws.cell(start_day+j,5+i).value = select_item
# A3セルに宛名の書き込み
# ws.cell(33, 4).value = "test"

# ファイル名を指定してブックを保存
wb.save("2020年度_健康管理チェックシート2021版.xlsx")